import sqlite3 as sq

import openpyxl

db = sq.connect('baza.db')
cursor = db.cursor()


async def db_start():
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS User (
            user_id TEXT PRIMARY KEY,
            username TEXT,
            firstname TEXT,
            lastname TEXT,
            created_at DATETIME,
            uz_balance REAL DEFAULT 0.0,
            btc_balance REAL DEFAULT 0.0,
            eth_balance REAL DEFAULT 0.0,
            usdt_balance REAL DEFAULT 0.0,
            bnb_balance REAL DEFAULT 0.0,
            sol_balance REAL DEFAULT 0.0,
            usdc_balance REAL DEFAULT 0.0,
            xrp_balance REAL DEFAULT 0.0,
            doge_balance REAL DEFAULT 0.0,
            ton_balance REAL DEFAULT 0.0,
            ada_balance REAL DEFAULT 0.0,
            is_bot TEXT
        )
    """)
    db.commit()


async def create_user(user_id, username, firstname, lastname, created_at, is_bot, uz_balance, btc_balance, eth_balance,
                      usdt_balance, bnb_balance, sol_balance, usdc_balance, xrp_balance, doge_balance, ton_balance,
                      ada_balance):
    user = cursor.execute("SELECT 1 FROM User WHERE user_id=='{key}'".format(key=user_id)).fetchone()
    if not user:
        cursor.execute("INSERT INTO User VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                       (
                           user_id, username, firstname, lastname, created_at, is_bot, uz_balance, btc_balance,
                           eth_balance,
                           usdt_balance, bnb_balance, sol_balance, usdc_balance, xrp_balance, doge_balance, ton_balance,
                           ada_balance))
        db.commit()


async def update_balance(user_id, valute, new_balance):
    column_name = f"{valute.lower()}_balance"

    cursor.execute("SELECT 1 FROM User WHERE user_id=?", (user_id,))
    user = cursor.fetchone()

    if user:
        cursor.execute(f"SELECT {column_name} FROM User WHERE user_id=?", (user_id,))
        current_balance = cursor.fetchone()[0]
        total_balance = float(current_balance) + float(new_balance)

        cursor.execute(f"UPDATE User SET {column_name}=? WHERE user_id=?", (total_balance, user_id))
        db.commit()


async def get_balance(user_id, valute):
    try:
        column_name = f"{valute.lower()}_balance"

        cursor.execute("SELECT {} FROM User WHERE user_id=?".format(column_name), (user_id,))
        balance = cursor.fetchone()

        if balance:
            return balance[0]
        else:
            return None

    except sq.Error as e:
        print(f"Error retrieving balance: {e}")
        return None


history_db = sq.connect('history.db')
history_cursor = history_db.cursor()


async def db_history():
    history_cursor.execute("""
            CREATE TABLE IF NOT EXISTS HistoryAction (
                user_id TEXT,
                username TEXT,
                firstname TEXT,
                lastname TEXT,
                created_at DATETIME,
                action_type TEXT,
                uz_sum REAL,
                crypto_name TEXT,
                crypto_sum REAL,
                card TEXT
            )
        """)
    history_db.commit()


def get_user_history(user_id):
    history_cursor.execute(
        "SELECT * FROM HistoryAction WHERE user_id = ?",
        (user_id,)
    )
    return history_cursor.fetchall()


async def insert_history(user_id, username, firstname, lastname, created_at, action_type, uz_sum, crypto_name,
                         crypto_sum, card):
    history_cursor.execute(
        "INSERT INTO HistoryAction (user_id, username, firstname, lastname, created_at, action_type, uz_sum, "
        "crypto_name, crypto_sum, card) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (user_id, username, firstname, lastname, created_at, action_type, uz_sum, crypto_name, crypto_sum, card)
    )
    history_db.commit()


def export_history_to_excel(user_id):
    history = get_user_history(user_id)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Monitoring"

    headers = ["User ID", "Username", "First Name", "Last Name", "Vaqt", "Faoliyat turi", "UZ Sum",
               "Kriptovalyuta", "Kriptovalyuta summasi", "Karta raqam"]

    for col_idx, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_idx, value=header)

    for row_idx, record in enumerate(history, start=2):
        for col_idx, field in enumerate(record, start=1):
            worksheet.cell(row=row_idx, column=col_idx, value=field)

    filename = f"{user_id}.xlsx"

    workbook.save(filename)

    return filename
